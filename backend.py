from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, send_from_directory, session
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
from datetime import datetime
import requests                                                                             
import urllib.parse
import os
import csv
import io
from io import StringIO, BytesIO

# ── IMPORT UNTUK EXCEL (Opsional, untuk export .xlsx) ─────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)

# Mengatasi bug Read-only file system di Vercel dengan memindahkan instance_path ke /tmp yang writable
if os.environ.get('VERCEL') or 'VERCEL' in os.environ:
    app.instance_path = '/tmp'

# Konfigurasi Database: Menggunakan PostgreSQL (Supabase) jika ada di environment variable, fallback ke SQLite untuk lokal
database_url = os.environ.get('DATABASE_URL') or os.environ.get('supabase_db') or os.environ.get('SUPABASE_DB')
if database_url:
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'

app.config['SECRET_KEY'] = 'your-secret-key-here'
db = SQLAlchemy(app)

# ── DECORATOR AUTH ADMIN ──────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ── KONFIGURASI ────────────────────────────────────────────────────────────────
DR_ASHOKA_WA = "6281230032017"  # Nomor WhatsApp Dr. Ashoka
ADMIN_EMAIL = "admin@email.com"  # Email admin untuk notifikasi

# ── MODEL DATABASE ────────────────────────────────────────────────────────────
class Response(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(100), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Jawaban 21 soal
    q1 = db.Column(db.Integer); q2 = db.Column(db.Integer); q3 = db.Column(db.Integer)
    q4 = db.Column(db.Integer); q5 = db.Column(db.Integer); q6 = db.Column(db.Integer)
    q7 = db.Column(db.Integer); q8 = db.Column(db.Integer); q9 = db.Column(db.Integer)
    q10 = db.Column(db.Integer); q11 = db.Column(db.Integer); q12 = db.Column(db.Integer)
    q13 = db.Column(db.Integer); q14 = db.Column(db.Integer); q15 = db.Column(db.Integer)
    q16 = db.Column(db.Integer); q17 = db.Column(db.Integer); q18 = db.Column(db.Integer)
    q19 = db.Column(db.Integer); q20 = db.Column(db.Integer); q21 = db.Column(db.Integer)
    
    # Skor
    skor_depresi = db.Column(db.Integer)
    skor_kecemasan = db.Column(db.Integer)
    skor_stres = db.Column(db.Integer)
    
    # Interpretasi
    ket_depresi = db.Column(db.String(20))
    ket_kecemasan = db.Column(db.String(20))
    ket_stres = db.Column(db.String(20))
    
    # Status
    status = db.Column(db.String(50))

# ── ATURAN SKORING DASS-21 ────────────────────────────────────────────────────
STRESS_ITEMS = [1, 6, 8, 11, 12, 14, 18]
ANXIETY_ITEMS = [2, 4, 7, 9, 15, 19, 20]
DEPRESSION_ITEMS = [3, 5, 10, 13, 16, 17, 21]

def get_interpretasi(skor):
    if skor <= 7: return "Normal"
    elif skor <= 9: return "Ringan"
    elif skor <= 12: return "Sedang"
    elif skor <= 16: return "Berat"
    else: return "Sangat Berat"

def hitung_skor(answers):
    dep = sum(answers[i-1] for i in DEPRESSION_ITEMS)
    anx = sum(answers[i-1] for i in ANXIETY_ITEMS)
    str_score = sum(answers[i-1] for i in STRESS_ITEMS)  
    return dep, anx, str_score

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/assets/<path:filename>')
def serve_assets(filename):
    return send_from_directory(os.path.join(app.root_path, 'assets'), filename)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin'))
        
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Validasi kredensial (Username: das, Password: kombinasi kuat)
        if username == 'das' and password == 'D@ss21_AdmLn!2026':
            session['admin_logged_in'] = True
            return redirect(url_for('admin'))
        else:
            error = 'Username atau password salah!'
            
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('login'))

@app.route('/admin')
@admin_required
def admin():
    responses = Response.query.order_by(Response.timestamp.desc()).all()
    return render_template('admin.html', responses=responses)

@app.route('/submit', methods=['POST'])
def submit():
    try:
        data = request.json
        nama = data.get('nama', 'Anonim')
        answers = [int(data.get(f'q{i}', 0)) for i in range(1, 22)]
        
        dep, anx, str_score = hitung_skor(answers)
        interp_dep = get_interpretasi(dep)
        interp_anx = get_interpretasi(anx)
        interp_str = get_interpretasi(str_score)
        
        is_critical = interp_dep in ['Berat', 'Sangat Berat'] or \
                      interp_anx in ['Berat', 'Sangat Berat'] or \
                      interp_str in ['Berat', 'Sangat Berat']
        
        status = "⚠️ BUTUH BANTUAN" if is_critical else "Aman"
        
        # Simpan ke database
        response = Response(
            nama=nama,
            q1=answers[0], q2=answers[1], q3=answers[2], q4=answers[3], q5=answers[4],
            q6=answers[5], q7=answers[6], q8=answers[7], q9=answers[8], q10=answers[9],
            q11=answers[10], q12=answers[11], q13=answers[12], q14=answers[13], q15=answers[14],
            q16=answers[15], q17=answers[16], q18=answers[17], q19=answers[18], q20=answers[19],
            q21=answers[20],
            skor_depresi=dep, skor_kecemasan=anx, skor_stres=str_score,
            ket_depresi=interp_dep, ket_kecemasan=interp_anx, ket_stres=interp_str,
            status=status
        )
        db.session.add(response)
        db.session.commit()
        
        if is_critical:
            send_whatsapp_alert(nama, dep, anx, str_score, interp_dep, interp_anx, interp_str)
        
        # ⚠️ KIRIM SEMUA DATA SKOR KE FRONTEND
        return jsonify({
            'success': True, 
            'message': 'Terima kasih telah mengisi!',
            'is_critical': is_critical,
            'status': status,
            'skor_depresi': dep,
            'skor_kecemasan': anx,
            'skor_stres': str_score,
            'interp_depresi': interp_dep,
            'interp_kecemasan': interp_anx,
            'interp_stres': interp_str,
            'kategori_tertinggi': "Depresi" if interp_dep in ['Berat', 'Sangat Berat'] else 
                                 ("Kecemasan" if interp_anx in ['Berat', 'Sangat Berat'] else 
                                 "Stres") if is_critical else "Kondisi Stabil"
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/delete/<int:id>', methods=['POST'])
@admin_required
def delete_response(id):
    response = Response.query.get_or_404(id)
    db.session.delete(response)
    db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/clear-all', methods=['POST'])
@admin_required
def clear_all():
    Response.query.delete()
    db.session.commit()
    return redirect(url_for('admin'))

# ── EXPORT KE CSV (EXCEL) - DIPERBAIKI ───────────────────────────────────────
@app.route('/admin/export')
@admin_required
def export_csv():
    responses = Response.query.order_by(Response.timestamp.desc()).all()
    
    # Buat CSV sebagai string dulu
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'ID', 'Waktu', 'Nama', 
        'Skor Depresi', 'Interpretasi Depresi',
        'Skor Kecemasan', 'Interpretasi Kecemasan',
        'Skor Stres', 'Interpretasi Stres',
        'Status'
    ])
    
    # Data
    for r in responses:
        writer.writerow([
            r.id,
            r.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            r.nama,
            r.skor_depresi, r.ket_depresi,
            r.skor_kecemasan, r.ket_kecemasan,
            r.skor_stres, r.ket_stres,
            r.status
        ])
    
    # Convert string to bytes (INI YANG DIPERBAIKI)
    csv_data = output.getvalue().encode('utf-8')
    output.close()
    
    # Create BytesIO from bytes
    output_bytes = BytesIO(csv_data)
    
    return send_file(
        output_bytes,
        mimetype='text/csv',
        as_attachment=True,
        download_name='dass21_results.csv'
    )

# ── EXPORT KE EXCEL (.XLSX) ─────────────────────────────────────────────────
@app.route('/admin/export-excel')
@admin_required
def export_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Library openpyxl tidak terinstall. Jalankan: pip install openpyxl'}), 500
    
    responses = Response.query.order_by(Response.timestamp.desc()).all()
    
    # Buat Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil DASS-21"
    
    # Header dengan styling
    headers = [
        'ID', 'Waktu', 'Nama', 
        'Skor Depresi', 'Ket. Depresi',
        'Skor Kecemasan', 'Ket. Kecemasan',
        'Skor Stres', 'Ket. Stres',
        'Status'
    ]
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, r in enumerate(responses, 2):
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=3, value=r.nama)
        ws.cell(row=row_idx, column=4, value=r.skor_depresi)
        ws.cell(row=row_idx, column=5, value=r.ket_depresi)
        ws.cell(row=row_idx, column=6, value=r.skor_kecemasan)
        ws.cell(row=row_idx, column=7, value=r.ket_kecemasan)
        ws.cell(row=row_idx, column=8, value=r.skor_stres)
        ws.cell(row=row_idx, column=9, value=r.ket_stres)
        ws.cell(row=row_idx, column=10, value=r.status)
        
        # Warna baris berdasarkan status
        if 'BUTUH' in r.status:
            fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
        else:
            fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).fill = fill
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='dass21_results.xlsx'
    )

# ── FUNGSI KIRIM WHATSAPP ALERT ───────────────────────────────────────────────
def send_whatsapp_alert(nama, dep, anx, str_score, i_dep, i_anx, i_str):
    max_score = max(dep, anx, str_score)
    kategori = "Depresi" if i_dep in ['Berat', 'Sangat Berat'] else \
               "Kecemasan" if i_anx in ['Berat', 'Sangat Berat'] else "Stres"
    
    pesan = f"Halo dr. Ashoka Sulistyasmara, ada pasien a.n {nama} yang hasil skrining DASS-21 nya {kategori} (Skor: {max_score}). Mohon tindak lanjutnya."
    link_wa = f"https://wa.me/{DR_ASHOKA_WA}?text={urllib.parse.quote(pesan)}"
    
    print(f"🚨 ALERT: {link_wa}")

# ── INISIALISASI DATABASE ─────────────────────────────────────────────────────
with app.app_context():
    db.create_all()

import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)